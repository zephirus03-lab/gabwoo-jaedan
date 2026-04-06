#!/usr/bin/env python3
"""
갑우문화사 재단지시서 엑셀 → JSON 변환기

메인 시트(재단지시서)에서 최신 날짜 섹션(진행 중)과
직전 섹션과 비교하여 완료된 작업을 추출합니다.

시트 구조:
- 시트1(재단지시서): 날짜별 섹션. 매일 갱신, 완료건은 다음 날 빠짐.
- 시트2,3: 관리자 별도 용도 (파싱 대상 아님)

사용법: python3 convert.py
입력: 생산진행현황/재단지시서_2026년.xlsx (읽기 전용)
출력: web/data.json
"""

import json
import os
import sys
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter


def find_sections(ws):
    """시트에서 모든 날짜 섹션의 (시작행, 날짜)를 반환"""
    sections = []
    for i in range(1, ws.max_row + 1):
        a = ws.cell(row=i, column=1).value
        if a and "설비별 재단 지시서" in str(a):
            s = ws.cell(row=i, column=19).value  # S열
            if isinstance(s, datetime):
                sections.append((i, s))
    return sections


def parse_section(ws, start_row, end_row, ws_fmt=None):
    """특정 섹션 범위의 작업 항목을 파싱.
    ws_fmt: 서식용 워크시트 (data_only=False로 연 것). C열 배경색으로 금일 작업 판별.
    """
    items = []
    current_equipment = ""
    data_start = start_row + 3  # 제목+메모+헤더 스킵

    for row_idx in range(data_start, end_row + 1):
        cells = {}
        for col in range(1, 19):
            cell = ws.cell(row=row_idx, column=col)
            try:
                letter = cell.column_letter
            except AttributeError:
                letter = get_column_letter(cell.column)
            cells[letter] = cell.value

        if not cells.get("B") and not cells.get("C"):
            continue

        c_val = str(cells.get("C", "") or "")
        b_val = str(cells.get("B", "") or "")
        if c_val.startswith("하기품목") or c_val.startswith("↑↑↑"):
            continue
        if b_val.startswith("↑↑↑") or "필히" in b_val:
            continue
        if cells.get("B") == "거래처":
            continue
        if not cells.get("B"):
            continue

        if cells.get("A"):
            a_val = str(cells["A"]).strip()
            if a_val not in ("설비", "설비별 재단 지시서") and "※" not in a_val:
                current_equipment = a_val

        # 납기
        deadline = cells.get("R")
        if isinstance(deadline, datetime):
            deadline_str = deadline.strftime("%Y-%m-%d")
        elif deadline and str(deadline).strip() not in ("", "-"):
            deadline_str = str(deadline).strip()
        else:
            deadline_str = "-"

        # 합계
        total = cells.get("J")
        if total is None:
            total_str = "-"
        elif isinstance(total, str):
            total_str = total.strip()
        else:
            total_str = str(int(total)) if total == int(total) else str(round(total, 2))

        # 통수
        quantity = cells.get("K")
        if quantity is None:
            quantity_str = "-"
        elif isinstance(quantity, (int, float)):
            quantity_str = f"{int(round(quantity)):,}"
        else:
            quantity_str = str(quantity)

        # 판수
        plates = cells.get("L")
        if plates is None:
            plates_str = "-"
        elif isinstance(plates, (int, float)):
            plates_str = str(int(plates))
        else:
            plates_str = str(plates)

        # 데이터 상태
        d_raw = str(cells.get("D", "") or "").strip()
        data_status = d_raw if d_raw else "-"

        # 금일 작업 진행 예정 판별: C열(3)~J열(10) 중 하나라도 theme=5 배경색이면
        is_today_work = False
        if ws_fmt:
            for check_col in range(3, 11):  # C~J열
                fmt_cell = ws_fmt.cell(row=row_idx, column=check_col)
                fill = fmt_cell.fill
                if fill.patternType == "solid" and fill.fgColor and fill.fgColor.theme == 5:
                    is_today_work = True
                    break

        item = {
            "row": row_idx,
            "equipment": current_equipment,
            "client": str(cells.get("B", "")).strip(),
            "product": str(cells.get("C", "")).strip(),
            "data_status": data_status,
            "sample": str(cells.get("E", "") or "").strip() or "-",
            "paper": str(cells.get("F", "") or "").strip() or "-",
            "size": str(cells.get("G", "") or "").strip() or "-",
            "cuts": str(cells.get("H", "") or "").strip() if cells.get("H") else "-",
            "grain": str(cells.get("I", "") or "").strip() or "-",
            "total": total_str,
            "quantity": quantity_str,
            "plates": plates_str,
            "method": str(cells.get("M", "") or "").strip() or "-",
            "category": str(cells.get("N", "") or "").strip() or "-",
            "plate_output": str(cells.get("O", "") or "").strip() or "-",
            "colors": str(cells.get("P", "") or "").strip() if cells.get("P") else "-",
            "post_process": str(cells.get("Q", "") or "").strip() or "-",
            "deadline": deadline_str,
            "is_today_work": is_today_work,
        }
        items.append(item)

    return items


def main():
    xlsx_path = "생산진행현황/재단지시서_2026년.xlsx"

    if not os.path.exists(xlsx_path):
        print(f"오류: 파일을 찾을 수 없습니다 - {xlsx_path}")
        sys.exit(1)

    print(f"엑셀 파일 읽는 중: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["재단지시서"]

    # 서식 정보용 (배경색 판별)
    print("  서식 정보 읽는 중...")
    wb_fmt = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws_fmt = wb_fmt["재단지시서"]

    # 모든 섹션 찾기
    sections = find_sections(ws)
    if len(sections) < 2:
        print("오류: 비교할 섹션이 부족합니다 (최소 2개 필요)")
        sys.exit(1)

    # 각 섹션의 끝 행 계산
    def section_end(idx):
        if idx + 1 < len(sections):
            return sections[idx + 1][0] - 1
        return ws.max_row

    # 오늘 이하 중 최신 섹션 찾기
    today = datetime.now().date()
    valid = [(i, r, d) for i, (r, d) in enumerate(sections) if d.date() <= today]
    if not valid:
        valid = [(i, r, d) for i, (r, d) in enumerate(sections)]

    latest_idx = max(valid, key=lambda x: x[2])[0]
    latest_date = sections[latest_idx][1].strftime("%Y-%m-%d")

    # 최신 섹션 파싱 (현재 진행 중) - 서식 정보 포함
    current_items = parse_section(ws, sections[latest_idx][0], section_end(latest_idx), ws_fmt)
    print(f"  최신 섹션: {latest_date} → 진행 중 {len(current_items)}건")

    # 완료 건 추출: 전체 섹션 연쇄 비교
    current_keys = {(i["client"], i["product"]) for i in current_items}

    def extract_completed(max_steps):
        """latest부터 역순으로 max_steps개 섹션을 비교하여 완료 건 추출"""
        items = []
        seen = set()
        for step in range(min(max_steps, latest_idx)):
            newer_idx = latest_idx - step
            older_idx = newer_idx - 1
            if older_idx < 0:
                break

            newer_date = sections[newer_idx][1].strftime("%Y-%m-%d")
            older_date = sections[older_idx][1].strftime("%Y-%m-%d")

            older_items = parse_section(ws, sections[older_idx][0], section_end(older_idx))
            newer_items_parsed = parse_section(ws, sections[newer_idx][0], section_end(newer_idx))
            newer_keys = {(i["client"], i["product"]) for i in newer_items_parsed}

            for item in older_items:
                key = (item["client"], item["product"])
                if key not in newer_keys and key not in current_keys and key not in seen:
                    item["completed_between"] = f"{older_date} → {newer_date}"
                    item["completed_date"] = newer_date
                    items.append(item)
                    seen.add(key)
        return items

    # 최근 10일 완료: 날짜 기준으로 10일 이내 섹션만
    latest_dt = sections[latest_idx][1]
    lookback_10d = 0
    for step in range(latest_idx):
        idx = latest_idx - step - 1
        if idx < 0:
            break
        delta = (latest_dt - sections[idx][1]).days
        if delta > 10:
            break
        lookback_10d = step + 1

    recent_completed = extract_completed(lookback_10d)
    recent_completed.sort(key=lambda x: x["client"])
    recent_completed.sort(key=lambda x: x["completed_date"], reverse=True)
    print(f"  최근 10일 완료: {len(recent_completed)}건 ({lookback_10d}개 섹션 비교)")

    # 전체 완료 (2026년): 모든 섹션 비교
    all_completed = extract_completed(latest_idx)
    all_completed.sort(key=lambda x: x["client"])
    all_completed.sort(key=lambda x: x["completed_date"], reverse=True)
    print(f"  전체 완료(2026년): {len(all_completed)}건 ({latest_idx}개 섹션 비교)")

    # 금일 작업 / 작업 대기 분리
    today_work = [i for i in current_items if i["is_today_work"]]
    waiting = [i for i in current_items if not i["is_today_work"]]

    print(f"  금일 작업 진행 예정: {len(today_work)}건")
    print(f"  작업 대기: {len(waiting)}건")

    # 거래처별 통계
    clients_all = {}
    for item in current_items:
        c = item["client"]
        clients_all[c] = clients_all.get(c, 0) + 1

    # 완료 건의 날짜별 건수
    recent_by_date = {}
    for item in recent_completed:
        d = item["completed_date"]
        recent_by_date[d] = recent_by_date.get(d, 0) + 1

    all_by_date = {}
    for item in all_completed:
        d = item["completed_date"]
        all_by_date[d] = all_by_date.get(d, 0) + 1

    # JSON 구성
    output = {
        "updated_at": latest_date,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "summary": {
            "today_work": len(today_work),
            "waiting": len(waiting),
            "recent_completed": len(recent_completed),
            "all_completed": len(all_completed),
            "total_clients": len(clients_all),
            "clients_in_progress": dict(sorted(clients_all.items(), key=lambda x: -x[1])),
            "recent_completed_by_date": dict(sorted(recent_by_date.items(), reverse=True)),
            "all_completed_by_date": dict(sorted(all_by_date.items(), reverse=True)),
        },
        "today_work": today_work,
        "waiting": waiting,
        "recent_completed": recent_completed,
        "all_completed": all_completed,
    }

    output_path = "web/data.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n변환 완료: {output_path}")
    print(f"  금일 작업 예정: {len(today_work)}건, 대기: {len(waiting)}건")
    print(f"  최근 10일 완료: {len(recent_completed)}건, 전체 완료: {len(all_completed)}건")
    print(f"  거래처: {len(clients_all)}개")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
갑우문화사 패키지사업부 생산현황 엑셀 → JSON 변환기.

입력 (읽기 전용, 절대 수정 금지):
  생산진행현황/수주관리 2026년도.xlsx
  생산진행현황/인쇄종합_*.xlsx   (glob)
  생산진행현황/출고품목리스트 (4월).xlsx

출력:
  web/data_package.json  (원자적 쓰기: tmp → os.replace)

설계 문서: docs/02-design/data-conversion.md
사용법:    python3 convert_package.py
"""

from __future__ import annotations

import glob
import json
import os
import re
import sys
import traceback
from datetime import datetime, date, timedelta
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# 경로 설정
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
SRC_DIR = BASE_DIR / "생산진행현황_패키지"
OUT_PATH = BASE_DIR / "web" / "data_package.json"

ORDERS_XLSX = SRC_DIR / "수주관리 2026년도.xlsx"
PRINTINGS_GLOB = str(SRC_DIR / "인쇄종합_*.xlsx")
SHIPMENTS_GLOB = str(SRC_DIR / "출고품목리스트 (*월).xlsx")

# 7개 표준 스테이지 순서 (UI 고정)
STAGES = ["인쇄", "코팅", "실크", "박", "형압", "톰슨", "접착"]

# 공정 키워드 매핑 (소문자 비교)
STAGE_KEYWORDS = {
    "인쇄": ["인쇄", "옵셋", "uv인쇄"],
    "코팅": ["코팅", "라미", "ir", "무광", "유광"],
    "실크": ["실크", "부분uv", "부분 uv"],
    "박":   ["박", "먹박", "금박", "홀로박"],  # "박스" 오탐지 주의 → 아래서 처리
    "형압": ["형압", "양각", "음각", "엠보"],
    "톰슨": ["톰슨", "타발"],
    "접착": ["접착", "이면접착", "삼면접착", "단면접착", "날개접착"],
}


# ---------------------------------------------------------------------------
# 공용 유틸
# ---------------------------------------------------------------------------
def _to_date_str(v):
    """datetime/date → 'YYYY-MM-DD', 그 외 None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    # 문자열로 들어오는 경우도 있을 수 있음 — 파싱 시도
    s = str(v).strip()
    if not s:
        return None
    m = re.match(r"(\d{4})[-./](\d{1,2})[-./](\d{1,2})", s)
    if m:
        y, mo, d = (int(x) for x in m.groups())
        try:
            return date(y, mo, d).strftime("%Y-%m-%d")
        except ValueError:
            return None
    return None


def _to_int(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        try:
            return int(v)
        except (ValueError, OverflowError):
            return None
    s = str(v).strip().replace(",", "")
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _to_text(v):
    if v is None:
        return ""
    return str(v).strip()


def _normalize_product(s: str) -> str:
    if not s:
        return ""
    return s.replace(" ", "").replace("_", "").replace("\t", "").lower()


# ---------------------------------------------------------------------------
# 공정 매핑
# ---------------------------------------------------------------------------
def extract_progress(gongjeong_text: str, stage_dates: dict) -> dict:
    """공정 자유 텍스트 → applicable_stages + progress(7 keys)."""
    text = (gongjeong_text or "").lower()
    applicable = []
    for stage in STAGES:
        hit = False
        for kw in STAGE_KEYWORDS[stage]:
            kw_l = kw.lower()
            if stage == "박":
                # "박스" 오탐지 방지: "박" 단독 키워드는 "박스" 제거 후 검사
                if kw_l == "박":
                    cleaned = text.replace("박스", "")
                    if "박" in cleaned:
                        hit = True
                        break
                else:
                    if kw_l in text:
                        hit = True
                        break
            else:
                if kw_l in text:
                    hit = True
                    break
        if hit:
            applicable.append(stage)

    if not applicable:
        # fallback: 전체 스테이지
        applicable = list(STAGES)

    progress = {
        stage: _to_date_str(stage_dates.get(stage)) for stage in STAGES
    }
    return {"progress": progress, "applicable_stages": applicable}


# ---------------------------------------------------------------------------
# 수주관리 파서
# ---------------------------------------------------------------------------
def _detect_header_row(sheet_rows: list) -> int:
    """'관리번호'가 등장하는 첫 행 인덱스 반환. 실패 시 예외."""
    for i, row in enumerate(sheet_rows):
        if row is None:
            continue
        for cell in row:
            if cell is not None and "관리번호" in str(cell):
                return i
    raise ValueError("헤더(관리번호) 탐지 실패")


def parse_orders(filepath: str) -> list:
    if not os.path.exists(filepath):
        raise FileNotFoundError(filepath)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        sheet_re = re.compile(r"수주관리\s*2026\.\s*(\d{1,2})월")
        monthly = []
        for name in wb.sheetnames:
            m = sheet_re.search(name)
            if m:
                monthly.append((int(m.group(1)), name))
        monthly.sort(key=lambda x: x[0])
        if not monthly:
            print("WARN: 수주관리 월별 시트를 찾지 못함")
            return []

        all_orders = {}
        for month_num, sheet_name in monthly:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            try:
                header_idx = _detect_header_row(rows)
            except ValueError:
                print(f"WARN: {sheet_name} 헤더 탐지 실패, skip")
                continue

            source_month = f"2026.{month_num:02d}"
            parsed = 0
            empty_streak = 0
            for row in rows[header_idx + 1:]:
                if row is None:
                    empty_streak += 1
                    if empty_streak >= 3:
                        break
                    continue
                # col 0 = No. None이면 종료 조건 (연속 빈 행 허용)
                if not row or row[0] is None:
                    empty_streak += 1
                    if empty_streak >= 3:
                        break
                    continue
                empty_streak = 0

                # 최소 20개 컬럼 확보
                cells = list(row) + [None] * (20 - len(row)) if len(row) < 20 else list(row)

                mgmt_no = _to_text(cells[3])
                if not mgmt_no or mgmt_no == "관리번호":
                    continue
                client = _to_text(cells[4])
                product = _to_text(cells[5])
                if not client and not product:
                    continue

                stage_dates = {
                    "인쇄": cells[10],
                    "코팅": cells[11],
                    "실크": cells[12],
                    "박":   cells[13],
                    "형압": cells[14],
                    "톰슨": cells[15],
                    "접착": cells[16],
                }
                process_spec = _to_text(cells[9])
                prog = extract_progress(process_spec, stage_dates)

                order = {
                    "mgmt_no": mgmt_no,
                    "manager": _to_text(cells[1]),
                    "registered_at": _to_date_str(cells[2]),
                    "client": client,
                    "product": product,
                    "qty": _to_int(cells[6]),
                    "net_qty": _to_int(cells[7]),
                    "deadline": _to_date_str(cells[8]),
                    "process_spec": process_spec,
                    "applicable_stages": prog["applicable_stages"],
                    "progress": prog["progress"],
                    "completed_at": _to_date_str(cells[17]),
                    "rework": _to_text(cells[18]) or None,
                    "note": _to_text(cells[19]),
                    "source_month": source_month,
                    "badges": {"printing_today": False, "shipping_today": False},
                }

                if mgmt_no in all_orders:
                    print(f"WARN: duplicate mgmt_no {mgmt_no} "
                          f"({all_orders[mgmt_no]['source_month']} → {source_month})")
                all_orders[mgmt_no] = order
                parsed += 1

            print(f"  {sheet_name}: {parsed}건")

        return list(all_orders.values())
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# 인쇄종합 파서
# ---------------------------------------------------------------------------
def parse_printings_today(filepath: str) -> list:
    if not os.path.exists(filepath):
        raise FileNotFoundError(filepath)

    today = datetime.now().date()
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        if "인쇄일정관리" not in wb.sheetnames:
            print("WARN: '인쇄일정관리' 시트 없음")
            return []
        ws = wb["인쇄일정관리"]

        header_map = {}
        header_row_idx = None
        data_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if header_row_idx is None:
                # 헤더 탐지: 행에 '계획일' 포함
                if row and any(
                    c is not None and "계획일" in str(c) for c in row
                ):
                    header_row_idx = i
                    for j, cell in enumerate(row):
                        if cell is None:
                            continue
                        key = str(cell).strip()
                        if key and key not in header_map:
                            header_map[key] = j
                continue
            data_rows.append(row)

        if header_row_idx is None:
            print("WARN: 인쇄일정관리 헤더 탐지 실패")
            return []

        def idx(*names):
            for n in names:
                if n in header_map:
                    return header_map[n]
            return None

        i_plan = idx("계획일")
        i_day = idx("Day", "요일")
        i_manager = idx("영업자", "담당")
        i_mgmt = idx("수주번호", "관리번호")
        i_client = idx("매출처", "거래처")
        i_process = idx("공정")
        i_product = idx("제품명", "품명")
        i_detail = idx("공정상세")
        i_deadline = idx("납품예정", "납기예정", "납기")
        i_qty = idx("수주량")
        i_sets = idx("벌수")
        i_net = idx("정미량")
        i_note = idx("비고")

        out = []
        for row in data_rows:
            if row is None or i_plan is None:
                continue
            if i_plan >= len(row):
                continue
            plan_val = row[i_plan]
            plan_date = None
            if isinstance(plan_val, datetime):
                plan_date = plan_val.date()
            elif isinstance(plan_val, date):
                plan_date = plan_val
            if plan_date != today:
                continue

            def get(idx_):
                if idx_ is None or idx_ >= len(row):
                    return None
                return row[idx_]

            out.append({
                "plan_date": _to_date_str(plan_val),
                "day": _to_text(get(i_day)) or None,
                "manager": _to_text(get(i_manager)),
                "mgmt_no": _to_text(get(i_mgmt)),
                "client": _to_text(get(i_client)),
                "process": _to_text(get(i_process)),
                "process_detail": _to_text(get(i_detail)),
                "product": _to_text(get(i_product)),
                "deadline": _to_date_str(get(i_deadline)),
                "qty": _to_int(get(i_qty)),
                "sets": _to_int(get(i_sets)),
                "net_qty": _to_int(get(i_net)),
                "note": _to_text(get(i_note)),
            })
        return out
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# 출고품목리스트 파서
# ---------------------------------------------------------------------------
def parse_shipments_today(filepath: str) -> list:
    if not os.path.exists(filepath):
        raise FileNotFoundError(filepath)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        mmdd_sheets = [s for s in wb.sheetnames if re.fullmatch(r"\d{4}", s)]
        if not mmdd_sheets:
            print("WARN: 출고리스트에 MMDD 시트 없음")
            return []
        latest = max(mmdd_sheets, key=int)
        ws = wb[latest]
        rows = list(ws.iter_rows(values_only=True))

        # sheet_date 추출 시도 (row 2 영역)
        sheet_date_str = None
        for row in rows[:3]:
            if not row:
                continue
            for cell in row:
                if isinstance(cell, (datetime, date)):
                    sheet_date_str = _to_date_str(cell)
                    break
            if sheet_date_str:
                break
        if not sheet_date_str:
            # fallback: 시트명 MMDD + 올해
            try:
                mm = int(latest[:2])
                dd = int(latest[2:])
                sheet_date_str = date(datetime.now().year, mm, dd).strftime("%Y-%m-%d")
            except ValueError:
                sheet_date_str = None

        # 헤더 동적 탐지: '업체명' 또는 '제품명'이 등장하는 행
        # 헤더 셀에 공백이 섞일 수 있으므로 whitespace 제거 후 매칭
        def _normalize(s):
            return re.sub(r"\s+", "", str(s or ""))
        header_idx = None
        header_map = {}
        for i, row in enumerate(rows):
            if not row:
                continue
            row_norm = [_normalize(c) for c in row]
            if any("업체" in t for t in row_norm) and any("제품" in t or "품명" in t for t in row_norm):
                header_idx = i
                for j, t in enumerate(row_norm):
                    if t and t not in header_map:
                        header_map[t] = j
                break

        if header_idx is None:
            print(f"WARN: 출고리스트 {latest} 헤더 탐지 실패")
            return []

        def idx(*names):
            for n in names:
                for k, v in header_map.items():
                    if n in k:
                        return v
            return None

        i_client = idx("업체")
        i_product = idx("제품", "품명")
        i_qty = idx("수량", "출고수량")
        i_dest = idx("출고처", "납품처")
        i_status = idx("비고", "상태")

        out = []
        for row in rows[header_idx + 1:]:
            if not row:
                continue
            def get(idx_):
                if idx_ is None or idx_ >= len(row):
                    return None
                return row[idx_]
            client = _to_text(get(i_client))
            if not client:
                continue
            out.append({
                "sheet_date": sheet_date_str,
                "client": client,
                "product": _to_text(get(i_product)),
                "qty": _to_int(get(i_qty)),
                "destination": _to_text(get(i_dest)),
                "status": _to_text(get(i_status)),
            })
        return out
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# 배지 / 요약
# ---------------------------------------------------------------------------
def merge_badges(orders: list, printings: list, shipments: list) -> list:
    printing_mgmt = {p.get("mgmt_no") for p in printings if p.get("mgmt_no")}
    shipment_products = {
        _normalize_product(s.get("product", "")): s
        for s in shipments
        if s.get("product")
    }

    printing_hit = 0
    shipping_hit = 0
    for o in orders:
        if o["mgmt_no"] in printing_mgmt:
            o["badges"]["printing_today"] = True
            printing_hit += 1
        pnorm = _normalize_product(o.get("product", ""))
        if pnorm and pnorm in shipment_products:
            o["badges"]["shipping_today"] = True
            shipping_hit += 1
    print(f"배지 부착: printing_today={printing_hit}, shipping_today={shipping_hit}")
    return orders


def compute_summary(orders: list, printings: list, shipments: list) -> dict:
    today = datetime.now().date()
    in_progress = sum(1 for o in orders if o["completed_at"] is None)

    # 7일 차트: today-6..today
    chart = {}
    for k in range(6, -1, -1):
        day = today - timedelta(days=k)
        chart[day.strftime("%Y-%m-%d")] = 0

    for o in orders:
        ca = o["completed_at"]
        if ca and ca in chart:
            chart[ca] += 1

    completed_7d = sum(chart.values())

    # 거래처 내림차순
    clients = {}
    for o in orders:
        if o["completed_at"] is None:
            c = o["client"]
            if c:
                clients[c] = clients.get(c, 0) + 1
    clients_sorted = dict(sorted(clients.items(), key=lambda x: -x[1]))

    return {
        "in_progress": in_progress,
        "printing_today": len(printings),
        "shipping_today": len(shipments),
        "completed_7d": completed_7d,
        "recent_7d_chart": chart,
        "clients_in_progress": clients_sorted,
    }


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------
def _resolve_input_files():
    missing = []
    if not ORDERS_XLSX.exists():
        missing.append(str(ORDERS_XLSX))

    printing_cands = sorted(glob.glob(PRINTINGS_GLOB))
    if not printing_cands:
        missing.append(PRINTINGS_GLOB)
    shipment_cands = sorted(glob.glob(SHIPMENTS_GLOB))
    if not shipment_cands:
        missing.append(SHIPMENTS_GLOB)

    if missing:
        print("오류: 입력 파일을 찾을 수 없습니다:")
        for m in missing:
            print(f"  - {m}")
        print("web/data_package.json 은 변경하지 않습니다.")
        sys.exit(2)

    # 인쇄종합: 첫 번째 후보
    printings_path = printing_cands[0]
    # 출고리스트: 가장 최근 mtime
    shipments_path = max(shipment_cands, key=lambda p: os.path.getmtime(p))
    return printings_path, shipments_path


def main():
    print(f"=== 패키지 생산현황 변환 (base={BASE_DIR}) ===")
    printings_path, shipments_path = _resolve_input_files()
    print(f"수주관리:   {ORDERS_XLSX.name}")
    print(f"인쇄종합:   {os.path.basename(printings_path)}")
    print(f"출고리스트: {os.path.basename(shipments_path)}")
    print()

    try:
        print("[1/3] 수주관리 파싱...")
        orders = parse_orders(str(ORDERS_XLSX))
        print(f"  → 총 {len(orders)}건")

        print("[2/3] 인쇄일정관리 파싱...")
        printings = parse_printings_today(printings_path)
        print(f"  → 오늘 {len(printings)}건")

        print("[3/3] 출고리스트 파싱...")
        shipments = parse_shipments_today(shipments_path)
        print(f"  → 최신 시트 {len(shipments)}건")

        orders = merge_badges(orders, printings, shipments)
        summary = compute_summary(orders, printings, shipments)

        now = datetime.now()
        output = {
            "updated_at": now.strftime("%Y-%m-%d"),
            "generated_at": now.strftime("%Y-%m-%d %H:%M"),
            "summary": summary,
            "orders": orders,
            "printings_today": printings,
            "shipments_today": shipments,
        }

        OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = OUT_PATH.with_suffix(".json.tmp")
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2)
        os.replace(tmp_path, OUT_PATH)

        size_kb = OUT_PATH.stat().st_size / 1024
        print()
        print(f"요약: 진행중 {summary['in_progress']}, "
              f"완료(7일) {summary['completed_7d']}, "
              f"오늘인쇄 {summary['printing_today']}, "
              f"오늘출고 {summary['shipping_today']}")
        print(f"→ {OUT_PATH} ({size_kb:.1f} KB)")
    except SystemExit:
        raise
    except Exception:
        traceback.print_exc()
        print("\n오류 발생. web/data_package.json 은 변경하지 않았습니다.")
        sys.exit(2)


if __name__ == "__main__":
    main()
